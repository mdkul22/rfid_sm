__author__ = "MDK"
from openpyxl import load_workbook
from collections import OrderedDict
from time import sleep
import serial
class Checker():

    def __init__(self):
        self.ser = serial.Serial('/dev/ttyACM0', 9600)
        main_dict = {
                'Mayunk Kulkarni': '70C3F3D',
                'Adarsh Pattanayak': '8060F5D',
                'Param Malhotra': 'FE95A07',
                'krishna chaitanya': 'AE9FCAB',
                'Debjeet Choudhury': 'F0BEF3D',
                'Mukul Joshi': 'DEBAA07',
                'ManuKoushik': 'FCBDA4C',
                'siddharth': '5EF6A17',
                'Rajendra Kotha': '6EA845D',
                'Aparna Roy': '6E90A17',
                'dharmendra baruah': '3E8849D',
                'suraj m': 'C68B408',
                'sebanti dasgupta': '7E06A27',
                'Kuber Nandwani': 'BE6A97D',
                'Shivang Bharadwaj': '4E86A17',
                'Priyanka Singh': 'C0D1F3D',
                'Srikriti Dahagam': '0E11A27',
                'Sharat Chandra': '7053F6D',
                'Nidhish Kamath': '1EAFA17',
                'Lakshya Bevli': 'CEE1A17',
                'Ritesh Patra': 'FE9410D',
                'K S Viswanath': '56B4468',
                'P.Lohit Kumar Reddy': 'DECC0D7',
                'Rishabh Pahuja': '40E6E3D',
                'Chirag Bhati': '609EE5D',
                'Shivang Kedia': 'C01DE8D',
                'manoj bhat': '9371ECE',
                'Pranshu Shubham': 'F000E5D',
                'Prachi Sinha': 'B0FEF3D',
                'Ayush Sinha': 'D2DD36B',
        }
        self.mainD = OrderedDict(sorted(main_dict.items(), key=lambda t: t[0]))
        d = self.mainD # making init ordered dict for ease in writing
        self.wb = load_workbook('logbook.xlsx') # creating workbook
        self.ws = self.wb.create_sheet(0) # create new sheet
        self.ws.title = self.get_serial('date') # sheet date as date
        # creating initial template
        p = ['New Session', 'Name', 'RFID code', 'Time Entered', 'Time Left']

        for i in range(len(p)):
            c = self.ws.cell(row=1, column=i+1)
            c.value = p[i]

        count = 0
        for i in d.keys():
            count = count + 1
            c = self.ws.cell(row=count, column=2)
            c.value = i
            c = self.ws.cell(row=count, column=3)
            c.value = d[i]

        self.flag = [] # creating flag values to be used later
        for number in range(0, 50):
            self.flag.append(0)
        self.wb.save('details.xlsx')

    def get_serial(self, r):
        val = self.ser.readline()
        val = str(val)
        val.strip()
        val.replace(" ", "")
        rfid = val[0:7]
        date = val[8:15]
        time = val[15:]
        if r == 'rfid':
            return rfid
        if r == 'date':
            return date
        if r == 'time':
            return time

    def check_in_out(self):
        count = 0
        rfid = self.get_serial('rfid')
        time_entry = None
        time_entry = self.get_serial('time')
        if time_entry == None:
            for i in self.mainD.values():
                count = count+1
                if rfid == i:
                    if self.flag[count] <= 2:
                        if self.flag[count] == 0:
                            self.flag[count] = 1
                            c = self.ws.cell(row=count+1, column=4)
                            c.value = time_entry
                            time_entry = None

                        elif self.flag[count] == 1:
                            self.flag[count] = 2
                            c = self.ws.cell(row=count+1, column=5)
                            c.value = time_entry
                            time_entry = None

                        elif self.flag[count] == 2:
                            self.flag[count] = 2
                            time_entry = None

                        else:
                            time_entry = None
                else:
                    continue
        self.wb.save('details.xlsx')



if __name__ == '__main__':
    x = Checker()
    while True:
        x.check_in_out()
        sleep(0.2)